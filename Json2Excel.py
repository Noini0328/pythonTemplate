import json
import csv
import pandas as pd
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Border, Side

pd.set_option("display.max_columns", None)
pd.set_option("display.max_rows", None)
pd.set_option("display.width", None)

# sample4 X(Twitter) direct message archive
def ExcelFormatter(data):
    wb = load_workbook(data)
    ws = wb.active

    # カラムの幅を指定
    column_widths = {'A': 35, 'B': 23, 'C': 23, 'D': 23, 'E': 50, 'F': 20, 'G': 20, 'H': 20}  # 例として幅を指定
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # セルに罫線を追加
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(top=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))

    # セルのスタイルを設定
    header_fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")
    for col_num, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # 6～8列目を折り返し表示
    for col_num in range(6, 9):
        column_letter = chr(65 + col_num - 1)
        for row_num, cell in enumerate(ws[column_letter], start=2):
            cell.alignment = Alignment(wrap_text=True)

    # 印刷設定を調整
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.print_area = ws.dimensions  # すべての列を1ページに印刷

    # フィルタ行を追加
    ws.auto_filter.ref = ws.dimensions

    # 変更を保存
    wb.save(data)


def JsonGetColumnNames(data, prefix=''):
    column_names = set()

    if isinstance(data, dict):
        for key, value in data.items():
            new_prefix = f"{prefix}.{key}" if prefix else key
            column_names.add(new_prefix)
            if isinstance(value, (dict, list)):
                column_names.update(JsonGetColumnNames(value, new_prefix))
    elif isinstance(data, list):
        for item in data:
            column_names.update(JsonGetColumnNames(item, prefix))

    return column_names

 
def JsonLoad(input_file):
    with open(input_file, 'r', encoding='utf-8') as json_file:
        content = json_file.read()
        start_index = min(content.find('['), content.find('{'))  # [ と { のうち、先に出現する位置を選択
        if start_index != -1:
            remaining_content = content[start_index:]
        else:
            remaining_content = content

        try:
            json_data = json.loads(remaining_content)
        except Exception as e:
            print("JSONデータを読み込めませんでした。エラー:", e)

    return json_data


def main():
    input_file = 'python/template/test.js'
    input_file = "python/template/direct-messages.js"

    json_data = JsonLoad(input_file)
    #print(json_data)

    json_column = JsonGetColumnNames(json_data)
    print(json_column)


    # conversationIdとmessagesのデータを抽出してリストに保存
    data = []
    for item in json_data:
        conversation_id = item["dmConversation"]["conversationId"]
        messages = item["dmConversation"]["messages"]
        for message in messages:
            message_data = message["messageCreate"]
            data.append({
                "conversationId": conversation_id,
                #"id": message_data["id"],
                "createdAt(JST)": datetime.strptime(message_data["createdAt"], "%Y-%m-%dT%H:%M:%S.%fZ") + timedelta(hours=9),
                "senderId": message_data["senderId"],
                "recipientId": message_data["recipientId"],
                "text": message_data["text"],
                "mediaUrls": message_data["mediaUrls"],
                "reactions": message_data["reactions"],
                "urls": message_data["urls"]
            })

    # リストからDataFrameを作成
    df = pd.DataFrame(data)
    #print(df)

    #df_index = ["conversationId", "createdAt", "senderId", "recipientId", "text", "mediaUrls", "reactions", "urls"]
    df.to_csv("output.csv", index=False)
    df.to_excel("output.xlsx", index=False)
    ExcelFormatter("output.xlsx")

if __name__ == "__main__":
    main()